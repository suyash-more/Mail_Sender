from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email_validator import validate_email, EmailNotValidError
from string import Template
import smtplib
import getpass
import openpyxl
import os
import sys


# https://myaccount.google.com/lesssecureapps enable the protection fro here

def read_template(filename):
    with open(filename, 'r', encoding='utf-8') as template_file:
        template_file_content = template_file.read()
    return Template(template_file_content)


filename_info = input("Enter the XLSX File name : ")
msg_info = input("Enter the Message File name : ")
MY_ADDRESS = input("Enter the Sender Mail Address : ")
send_name = input("Enter Sender Name : ")
PASSWORD = getpass.getpass('Enter Sender Password : ')
subject = input("Enter the Subject : ")

workbook = openpyxl.load_workbook(f'{os.path.abspath(filename_info)}')
sheet = workbook.active

try:
    message_template = read_template(f'{os.path.abspath(msg_info)}')

except Exception as e:
    print("The Message File Was Not Found")
    print(e)
    sys.exit()

s = smtplib.SMTP(host='smtp.gmail.com', port=587)
s.starttls()
s.login(MY_ADDRESS, PASSWORD)

print("\nSMTP Server is UP..!!\n")

for i in range(1, sheet.max_row + 1):
    email = str(sheet.cell(row=i, column=2).value)
    name = str(sheet.cell(row=i, column=1).value)
    if email == "None":
        continue
    if name == "None":
        name = "Sir"
    if subject == "":
        subject = "Hey this is a Trial..!!"
    try:
        valid = validate_email(email)
        email = valid.email

    except EmailNotValidError as e:
        print(f"{email} is not VALID")
        continue

    message = message_template.safe_substitute(RECV_NAME=name.title(), SEND_NAME=send_name.title())

    msg = MIMEMultipart()
    msg['From'] = MY_ADDRESS
    msg['To'] = email
    msg['Subject'] = subject

    msg.attach(MIMEText(message, 'plain'))

    s.send_message(msg)

    print(f"Email sent To : {email}")

    del msg

print("\n\nDone! Quitting server")
s.quit()
